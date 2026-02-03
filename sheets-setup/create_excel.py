"""
Lexile Reading Text DB - Excel 파일 생성 스크립트
4개 시트: READING_TEXT_MASTER, CONFIG, COVERAGE_MATRIX, QUESTION_BANK
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date
import os

# ==================== 색상/스타일 정의 ====================

BLUE_FILL = PatternFill(start_color="4285F4", end_color="4285F4", fill_type="solid")
GREEN_FILL = PatternFill(start_color="34A853", end_color="34A853", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
GRAY_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def apply_header_style(ws, row, col_start, col_end, fill=BLUE_FILL):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def apply_border(ws, row, col_start, col_end):
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER

# ==================== 샘플 데이터 ====================

SAMPLE_DATA = [
    ["L700-NAR-050-001","700-900",750,"Middle School","중1","Narrative","A Rainy Morning",52,"Micro",
     "Jisoo woke up to the sound of rain tapping against her window. She grabbed her yellow umbrella and stepped outside. The streets were quiet, and puddles reflected the gray sky. A small cat sat under a bench, watching her with curious eyes. Jisoo smiled and walked to school.",
     5,10.4,"A2/B1","워밍업","2026-02-01","Micro narrative"],
    ["L700-NAR-100-001","700-900",780,"Middle School","중1","Narrative","The New Student",103,"Short",
     "A new student named Marco arrived at Minjun's school on Monday morning. He came from Brazil and spoke very little Korean. During lunch, Minjun noticed Marco sitting alone in the cafeteria. He walked over and sat beside him. At first, they used simple English words and hand gestures to communicate. Marco showed Minjun pictures of his hometown on his phone. The colorful buildings and wide beaches looked amazing. By the end of the week, Marco had joined Minjun's friend group. They discovered that friendship doesn't always need perfect words — sometimes a smile is enough.",
     10,10.3,"A2/B1","수업","2026-02-01","Short narrative"],
    ["L700-NAR-200-001","700-900",820,"Middle School","중2","Narrative","The Lost Notebook",198,"Medium",
     "Soyeon had always kept a notebook filled with her drawings and private thoughts. She carried it everywhere — to school, to the park, and even to family dinners. One Thursday afternoon, she realized the notebook was gone. She searched her backpack three times, checked under her desk, and asked every classmate she could find. Nobody had seen it. That evening, Soyeon could barely eat dinner. Her mother noticed her quiet mood and asked what was wrong. Soyeon explained about the lost notebook, expecting her mother to say it was just a notebook. Instead, her mother said something surprising. She told Soyeon that when she was young, she had lost her diary too, and it felt like losing a piece of herself. The next morning, Soyeon found the notebook sitting on her desk at school. A sticky note on the cover read: I found this in the library. Your drawings are beautiful. Soyeon never discovered who returned it, but she felt grateful. She realized that even strangers sometimes care about the things that matter to us.",
     17,11.6,"A2/B1","수업","2026-02-01","Medium narrative"],
    ["L700-EXP-050-001","700-900",760,"Middle School","중1","Expository","Why We Sleep",48,"Micro",
     "Sleep is essential for both the body and the mind. During sleep, the brain organizes memories and repairs damaged cells. Without enough rest, people have trouble concentrating and making decisions. Most teenagers need about eight to ten hours of sleep each night.",
     4,12.0,"A2/B1","워밍업","2026-02-01","Micro expository"],
    ["L700-EXP-100-001","700-900",800,"Middle School","중2","Expository","How Volcanoes Form",105,"Short",
     "Volcanoes form when hot melted rock called magma rises from deep inside the Earth. The Earth's outer layer is made up of large pieces called tectonic plates. These plates are constantly moving, although very slowly. When two plates push against each other or pull apart, cracks form in the surface. Magma pushes through these cracks and reaches the surface as lava. Over time, layers of cooled lava build up and create the cone shape that most people picture when they think of a volcano. Some volcanoes erupt violently with ash and fire, while others release lava slowly and quietly. Scientists study volcanoes carefully to predict eruptions and protect nearby communities.",
     10,10.5,"A2/B1","수업","2026-02-01","Short expository"],
    ["L700-INF-100-001","700-900",790,"Middle School","중1","Informational","Recycling in South Korea",98,"Short",
     "South Korea is one of the world's leading countries in recycling. The government introduced strict recycling rules in the 1990s that require citizens to separate their waste into several categories. Most neighborhoods have designated areas where residents must sort their trash into food waste, plastics, paper, glass, and general waste. The food waste is often turned into animal feed or fertilizer. As a result of these efforts, South Korea now recycles more than half of its total waste. Many other countries are studying Korea's system as a model for improving their own recycling programs.",
     8,12.3,"A2/B1","수업","2026-02-01","Short informational"],
    ["L700-ARG-200-001","700-900",860,"Middle School","중2","Argumentative","School Uniforms",196,"Medium",
     "Many students complain about having to wear school uniforms every day. They argue that uniforms limit their freedom of expression and make everyone look the same. While these feelings are understandable, school uniforms actually provide several important benefits that students may not immediately recognize. First, uniforms create a sense of equality among students. When everyone wears the same clothes, it becomes harder to judge others based on their family's income or fashion choices. This can reduce bullying and social pressure related to appearance. Second, uniforms save time and reduce stress in the morning. Students do not need to spend time deciding what to wear, which means they can focus on preparing for their classes instead. Third, research from several countries suggests that schools with uniform policies tend to have fewer discipline problems and better academic performance. Of course, students should have opportunities to express their individuality through other means, such as art, sports, and extracurricular activities. However, the evidence suggests that the advantages of school uniforms outweigh the disadvantages. Rather than viewing uniforms as a restriction, students might consider them as a tool that supports a more focused and fair learning environment.",
     16,12.3,"B1","수업","2026-02-01","Medium argumentative"],
    ["L700-PRO-050-001","700-900",730,"Middle School","중1","Procedural","Making a Paper Airplane",50,"Micro",
     "To make a simple paper airplane, first fold a sheet of paper in half lengthwise. Then unfold it and fold the top two corners down to the center line. Fold the angled edges to the center again. Finally, fold the paper in half and create two wings by folding each side down.",
     4,12.5,"A2","워밍업","2026-02-01","Micro procedural"],
    ["L700-LIT-050-001","700-900",740,"Middle School","중1","Literary","Morning Light",46,"Micro",
     "The morning sun crept through the curtains like a shy visitor. Dust particles danced in the golden light, spinning slowly without purpose. For a moment, the room felt like a quiet painting — still and warm and perfectly at peace.",
     3,15.3,"A2/B1","워밍업","2026-02-01","Micro literary"],
    ["L700-LIT-100-001","700-900",810,"Middle School","중2","Literary","The Old Bench",101,"Short",
     "There is a wooden bench in the corner of our school playground that nobody seems to notice anymore. Its paint is peeling, and one leg is slightly shorter than the others, making it wobble when you sit down. But I love that bench. I sit there every afternoon and watch the clouds change shape above the basketball court. Sometimes I bring a book. Sometimes I just sit and think. My friends ask me why I don't join them in the classroom during break. I tell them that the bench listens better than most people do. It never interrupts, never judges, and never tells me to hurry up.",
     9,11.2,"A2/B1","수업","2026-02-01","Short literary"],
]

# ==================== 빈 템플릿 슬롯 ====================

TEMPLATE_SLOTS = [
    # 100-300L
    ["","100-300","","Early Elementary","초1","Narrative","",50,"Micro","","","","Pre-A1","워밍업","",""],
    ["","100-300","","Early Elementary","초1","Narrative","",100,"Short","","","","A1","수업","",""],
    ["","100-300","","Early Elementary","초2","Narrative","",200,"Medium","","","","A1","수업","",""],
    ["","100-300","","Early Elementary","초1","Expository","",50,"Micro","","","","Pre-A1","워밍업","",""],
    ["","100-300","","Early Elementary","초2","Expository","",100,"Short","","","","A1","수업","",""],
    ["","100-300","","Early Elementary","초1","Procedural","",50,"Micro","","","","Pre-A1","워밍업","",""],
    ["","100-300","","Early Elementary","초1","Procedural","",100,"Short","","","","A1","수업","",""],
    # 300-500L
    ["","300-500","","Upper Elementary","초3","Narrative","",50,"Micro","","","","A1/A2","워밍업","",""],
    ["","300-500","","Upper Elementary","초3","Narrative","",100,"Short","","","","A1/A2","수업","",""],
    ["","300-500","","Upper Elementary","초4","Narrative","",200,"Medium","","","","A2","수업","",""],
    ["","300-500","","Upper Elementary","초4","Narrative","",350,"Long","","","","A2","다독","",""],
    ["","300-500","","Upper Elementary","초3","Expository","",50,"Micro","","","","A1/A2","워밍업","",""],
    ["","300-500","","Upper Elementary","초4","Expository","",100,"Short","","","","A2","수업","",""],
    ["","300-500","","Upper Elementary","초4","Expository","",200,"Medium","","","","A2","수업","",""],
    ["","300-500","","Upper Elementary","초3","Informational","",100,"Short","","","","A1/A2","수업","",""],
    ["","300-500","","Upper Elementary","초3","Procedural","",50,"Micro","","","","A1/A2","워밍업","",""],
    ["","300-500","","Upper Elementary","초3","Procedural","",100,"Short","","","","A2","수업","",""],
    ["","300-500","","Upper Elementary","초4","Literary","",50,"Micro","","","","A2","워밍업","",""],
    # 500-700L
    ["","500-700","","Transitional","초5","Narrative","",50,"Micro","","","","A2/B1","워밍업","",""],
    ["","500-700","","Transitional","초5","Narrative","",100,"Short","","","","A2/B1","수업","",""],
    ["","500-700","","Transitional","초6","Narrative","",200,"Medium","","","","A2/B1","수업","",""],
    ["","500-700","","Transitional","초6","Narrative","",350,"Long","","","","B1","다독","",""],
    ["","500-700","","Transitional","초5","Expository","",50,"Micro","","","","A2","워밍업","",""],
    ["","500-700","","Transitional","초6","Expository","",100,"Short","","","","A2/B1","수업","",""],
    ["","500-700","","Transitional","초6","Expository","",200,"Medium","","","","B1","수업","",""],
    ["","500-700","","Transitional","초5","Informational","",100,"Short","","","","A2/B1","수업","",""],
    ["","500-700","","Transitional","초6","Informational","",200,"Medium","","","","B1","수업","",""],
    ["","500-700","","Transitional","초6","Argumentative","",200,"Medium","","","","B1","수업","",""],
    ["","500-700","","Transitional","초5","Procedural","",50,"Micro","","","","A2","워밍업","",""],
    ["","500-700","","Transitional","초5","Procedural","",100,"Short","","","","A2/B1","수업","",""],
    ["","500-700","","Transitional","초6","Literary","",100,"Short","","","","A2/B1","수업","",""],
    ["","500-700","","Transitional","초6","Literary","",200,"Medium","","","","B1","다독","",""],
    # 900-1100L
    ["","900-1100","","Upper Secondary","중3","Narrative","",50,"Micro","","","","B1","워밍업","",""],
    ["","900-1100","","Upper Secondary","중3","Narrative","",100,"Short","","","","B1/B2","수업","",""],
    ["","900-1100","","Upper Secondary","고1","Narrative","",200,"Medium","","","","B1/B2","수업","",""],
    ["","900-1100","","Upper Secondary","고1","Narrative","",350,"Long","","","","B2","다독","",""],
    ["","900-1100","","Upper Secondary","중3","Expository","",50,"Micro","","","","B1","워밍업","",""],
    ["","900-1100","","Upper Secondary","고1","Expository","",100,"Short","","","","B1/B2","수업","",""],
    ["","900-1100","","Upper Secondary","고1","Expository","",200,"Medium","","","","B2","수업","",""],
    ["","900-1100","","Upper Secondary","고1","Expository","",350,"Long","","","","B2","시험대비","",""],
    ["","900-1100","","Upper Secondary","중3","Informational","",100,"Short","","","","B1/B2","수업","",""],
    ["","900-1100","","Upper Secondary","고1","Informational","",200,"Medium","","","","B2","수업","",""],
    ["","900-1100","","Upper Secondary","고1","Informational","",350,"Long","","","","B2","시험대비","",""],
    ["","900-1100","","Upper Secondary","고1","Argumentative","",100,"Short","","","","B1/B2","수업","",""],
    ["","900-1100","","Upper Secondary","고1","Argumentative","",200,"Medium","","","","B2","수업","",""],
    ["","900-1100","","Upper Secondary","고1","Argumentative","",350,"Long","","","","B2","시험대비","",""],
    ["","900-1100","","Upper Secondary","중3","Procedural","",100,"Short","","","","B1","수업","",""],
    ["","900-1100","","Upper Secondary","고1","Literary","",100,"Short","","","","B1/B2","수업","",""],
    ["","900-1100","","Upper Secondary","고1","Literary","",200,"Medium","","","","B2","다독","",""],
    ["","900-1100","","Upper Secondary","고1","Literary","",350,"Long","","","","B2","시험대비","",""],
    # 1100-1300L
    ["","1100-1300","","Pre-CSAT","고2","Narrative","",100,"Short","","","","B2","수업","",""],
    ["","1100-1300","","Pre-CSAT","고2","Narrative","",200,"Medium","","","","B2","수업","",""],
    ["","1100-1300","","Pre-CSAT","고3","Narrative","",350,"Long","","","","B2/C1","시험대비","",""],
    ["","1100-1300","","Pre-CSAT","고2","Expository","",100,"Short","","","","B2","수업","",""],
    ["","1100-1300","","Pre-CSAT","고2","Expository","",200,"Medium","","","","B2","수업","",""],
    ["","1100-1300","","Pre-CSAT","고3","Expository","",350,"Long","","","","B2/C1","시험대비","",""],
    ["","1100-1300","","Pre-CSAT","고2","Informational","",100,"Short","","","","B2","수업","",""],
    ["","1100-1300","","Pre-CSAT","고3","Informational","",200,"Medium","","","","B2/C1","수업","",""],
    ["","1100-1300","","Pre-CSAT","고3","Informational","",350,"Long","","","","B2/C1","시험대비","",""],
    ["","1100-1300","","Pre-CSAT","고2","Argumentative","",100,"Short","","","","B2","수업","",""],
    ["","1100-1300","","Pre-CSAT","고3","Argumentative","",200,"Medium","","","","B2/C1","수업","",""],
    ["","1100-1300","","Pre-CSAT","고3","Argumentative","",350,"Long","","","","B2/C1","시험대비","",""],
    ["","1100-1300","","Pre-CSAT","고2","Procedural","",100,"Short","","","","B2","수업","",""],
    ["","1100-1300","","Pre-CSAT","고2","Literary","",100,"Short","","","","B2","수업","",""],
    ["","1100-1300","","Pre-CSAT","고3","Literary","",200,"Medium","","","","B2/C1","다독","",""],
    ["","1100-1300","","Pre-CSAT","고3","Literary","",350,"Long","","","","B2/C1","시험대비","",""],
    # 1300-1500L
    ["","1300-1500","","Academic","대학1","Expository","",200,"Medium","","","","C1","수업","",""],
    ["","1300-1500","","Academic","대학1","Expository","",350,"Long","","","","C1","과제","",""],
    ["","1300-1500","","Academic","대학1","Informational","",200,"Medium","","","","C1","수업","",""],
    ["","1300-1500","","Academic","대학1","Informational","",350,"Long","","","","C1","과제","",""],
    ["","1300-1500","","Academic","대학1","Argumentative","",200,"Medium","","","","C1","수업","",""],
    ["","1300-1500","","Academic","대학1","Argumentative","",350,"Long","","","","C1","과제","",""],
    ["","1300-1500","","Academic","대학1","Literary","",200,"Medium","","","","B2/C1","다독","",""],
    ["","1300-1500","","Academic","대학1","Literary","",350,"Long","","","","C1","과제","",""],
]


def create_master_sheet(wb):
    """READING_TEXT_MASTER 시트 생성"""
    ws = wb.active
    ws.title = "READING_TEXT_MASTER"

    headers = [
        "text_id", "lexile_band", "lexile_score", "age_group", "grade_hint",
        "genre", "topic", "word_count", "length_type", "text_body",
        "sentence_count", "avg_sentence_length", "vocabulary_band",
        "intended_use", "created_date", "notes"
    ]

    # 헤더
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    apply_header_style(ws, 1, 1, len(headers))
    ws.freeze_panes = "A2"

    # 샘플 데이터 (700-900L)
    for i, row_data in enumerate(SAMPLE_DATA, 2):
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = THIN_BORDER
            if j == 10:  # text_body
                cell.alignment = Alignment(wrap_text=True)
        # 샘플 행 배경색
        for j in range(1, len(headers) + 1):
            ws.cell(row=i, column=j).fill = LIGHT_GREEN_FILL

    # 빈 템플릿 슬롯
    start_row = len(SAMPLE_DATA) + 2
    current_band = None
    for i, slot in enumerate(TEMPLATE_SLOTS):
        row_num = start_row + i
        band = slot[1]

        for j, val in enumerate(slot, 1):
            cell = ws.cell(row=row_num, column=j, value=val if val != "" else None)
            cell.border = THIN_BORDER

        # Band 시작 색상
        if band != current_band:
            current_band = band
            for j in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=j).fill = YELLOW_FILL
        else:
            for j in range(1, len(headers) + 1):
                if ws.cell(row=row_num, column=j).fill == PatternFill():
                    pass  # keep default

    # 컬럼 너비
    widths = {
        1: 22, 2: 12, 3: 12, 4: 18, 5: 10,
        6: 16, 7: 20, 8: 12, 9: 12, 10: 80,
        11: 14, 12: 18, 13: 14, 14: 12, 15: 14, 16: 30
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # 데이터 검증 (드롭다운)
    band_list = '"100-300,300-500,500-700,700-900,900-1100,1100-1300,1300-1500"'
    genre_list = '"Narrative,Expository,Informational,Argumentative,Procedural,Literary"'
    length_list = '"Micro,Short,Medium,Long"'
    vocab_list = '"Pre-A1,A1,A1/A2,A2,A2/B1,B1,B1/B2,B2,B2/C1,C1"'
    use_list = '"수업,다독,과제,시험대비,워밍업"'

    last_data_row = start_row + len(TEMPLATE_SLOTS) + 100

    dv_band = DataValidation(type="list", formula1=band_list, allow_blank=True)
    dv_band.error = "유효한 Lexile Band를 선택하세요"
    dv_band.errorTitle = "잘못된 값"
    ws.add_data_validation(dv_band)
    dv_band.add(f"B2:B{last_data_row}")

    dv_genre = DataValidation(type="list", formula1=genre_list, allow_blank=True)
    ws.add_data_validation(dv_genre)
    dv_genre.add(f"F2:F{last_data_row}")

    dv_length = DataValidation(type="list", formula1=length_list, allow_blank=True)
    ws.add_data_validation(dv_length)
    dv_length.add(f"I2:I{last_data_row}")

    dv_vocab = DataValidation(type="list", formula1=vocab_list, allow_blank=True)
    ws.add_data_validation(dv_vocab)
    dv_vocab.add(f"M2:M{last_data_row}")

    dv_use = DataValidation(type="list", formula1=use_list, allow_blank=True)
    ws.add_data_validation(dv_use)
    dv_use.add(f"N2:N{last_data_row}")

    return ws


def create_config_sheet(wb):
    """CONFIG 시트 생성"""
    ws = wb.create_sheet("CONFIG")

    # ---- Lexile Bands (A열) ----
    ws.cell(row=1, column=1, value="Lexile Bands").font = TITLE_FONT
    sub_headers = ["Band", "Age Group", "Grade", "Purpose"]
    for i, h in enumerate(sub_headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = BOLD_FONT
        cell.fill = LIGHT_BLUE_FILL
        cell.border = THIN_BORDER

    bands = [
        ["100-300", "Early Elementary", "초1-2", "영어 문장 감각 형성"],
        ["300-500", "Upper Elementary", "초3-4", "의미 단위 읽기"],
        ["500-700", "Transitional", "초5-6", "문단 이해 시작"],
        ["700-900", "Middle School", "중1-2", "정보 이해"],
        ["900-1100", "Upper Secondary", "중3-고1", "논리·원인결과"],
        ["1100-1300", "Pre-CSAT", "고2-3", "추상 개념"],
        ["1300-1500", "Academic", "대학 초반", "학술 독해"],
    ]
    for i, row_data in enumerate(bands, 3):
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = THIN_BORDER

    # ---- Genres (F열) ----
    ws.cell(row=1, column=6, value="Genres").font = TITLE_FONT
    g_headers = ["Code", "Name", "Thinking Type"]
    for i, h in enumerate(g_headers):
        cell = ws.cell(row=2, column=6 + i, value=h)
        cell.font = BOLD_FONT
        cell.fill = LIGHT_BLUE_FILL
        cell.border = THIN_BORDER

    genres = [
        ["NAR", "Narrative", "시간·사건"],
        ["EXP", "Expository", "설명"],
        ["INF", "Informational", "사실"],
        ["ARG", "Argumentative", "주장·근거"],
        ["PRO", "Procedural", "과정"],
        ["LIT", "Literary", "정서·표현"],
    ]
    for i, row_data in enumerate(genres, 3):
        for j, val in enumerate(row_data):
            cell = ws.cell(row=i, column=6 + j, value=val)
            cell.border = THIN_BORDER

    # ---- Length Types (J열) ----
    ws.cell(row=1, column=10, value="Length Types").font = TITLE_FONT
    l_headers = ["Type", "Target Words", "Range", "Purpose"]
    for i, h in enumerate(l_headers):
        cell = ws.cell(row=2, column=10 + i, value=h)
        cell.font = BOLD_FONT
        cell.fill = LIGHT_BLUE_FILL
        cell.border = THIN_BORDER

    lengths = [
        ["Micro", 50, "40-60", "워밍업 / 문장 훈련"],
        ["Short", 100, "80-120", "수업용 핵심 독해"],
        ["Medium", 200, "170-230", "정독 훈련"],
        ["Long", 350, "280-420", "다독 / 시험 대비"],
    ]
    for i, row_data in enumerate(lengths, 3):
        for j, val in enumerate(row_data):
            cell = ws.cell(row=i, column=10 + j, value=val)
            cell.border = THIN_BORDER

    # ---- CEFR Bands (O열) ----
    ws.cell(row=1, column=15, value="CEFR Bands").font = TITLE_FONT
    ws.cell(row=2, column=15, value="Level").font = BOLD_FONT
    ws.cell(row=2, column=15).fill = LIGHT_BLUE_FILL
    ws.cell(row=2, column=15).border = THIN_BORDER
    cefr = ["Pre-A1", "A1", "A1/A2", "A2", "A2/B1", "B1", "B1/B2", "B2", "B2/C1", "C1"]
    for i, v in enumerate(cefr, 3):
        cell = ws.cell(row=i, column=15, value=v)
        cell.border = THIN_BORDER

    # ---- Intended Uses (P열) ----
    ws.cell(row=1, column=16, value="Intended Uses").font = TITLE_FONT
    ws.cell(row=2, column=16, value="Use").font = BOLD_FONT
    ws.cell(row=2, column=16).fill = LIGHT_BLUE_FILL
    ws.cell(row=2, column=16).border = THIN_BORDER
    uses = ["수업", "다독", "과제", "시험대비", "워밍업"]
    for i, v in enumerate(uses, 3):
        cell = ws.cell(row=i, column=16, value=v)
        cell.border = THIN_BORDER

    # ---- 연령 적합 주제 (A12~) ----
    ws.cell(row=12, column=1, value="연령대별 적합 주제").font = TITLE_FONT
    age_headers = ["Age Group", "허용 주제", "제한 주제"]
    for i, h in enumerate(age_headers, 1):
        cell = ws.cell(row=13, column=i, value=h)
        cell.font = BOLD_FONT
        cell.fill = LIGHT_BLUE_FILL
        cell.border = THIN_BORDER

    age_topics = [
        ["Early Elementary", "가족, 동물, 학교, 놀이, 음식", "정치, 추상 철학, 폭력"],
        ["Upper Elementary", "가족, 동물, 학교, 자연, 취미", "정치, 이념, 추상 철학"],
        ["Middle School", "관계, 사회 현상, 기술, 자연, 문화", "이념 논쟁, 성인 주제"],
        ["Upper Secondary", "환경, 기술, 윤리, 사회, 과학", "지나친 전문 논문"],
        ["Pre-CSAT", "환경, 기술, 윤리, 철학, 경제", "편향된 이념"],
        ["Academic", "학술, 비평, 연구, 전문분야", "거의 없음"],
    ]
    for i, row_data in enumerate(age_topics, 14):
        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.border = THIN_BORDER

    # 컬럼 너비
    for col in range(1, 17):
        ws.column_dimensions[get_column_letter(col)].width = 20

    return ws


def create_coverage_sheet(wb):
    """COVERAGE_MATRIX 시트 생성"""
    ws = wb.create_sheet("COVERAGE_MATRIX")

    ws.cell(row=1, column=1, value="COVERAGE MATRIX — Genre × Length Type").font = TITLE_FONT

    genres = ["Narrative", "Expository", "Informational", "Argumentative", "Procedural", "Literary"]
    lengths = ["Micro", "Short", "Medium", "Long"]
    all_bands = ["100-300", "300-500", "500-700", "700-900", "900-1100", "1100-1300", "1300-1500"]

    # 700-900 샘플 데이터 카운트
    sample_counts = {}
    for s in SAMPLE_DATA:
        genre = s[5]
        length_type = s[8]
        key = f"{genre}|{length_type}"
        sample_counts[key] = sample_counts.get(key, 0) + 1

    current_row = 3
    for band in all_bands:
        ws.cell(row=current_row, column=1, value=f"Lexile Band: {band}").font = SECTION_FONT
        ws.cell(row=current_row, column=1).fill = LIGHT_BLUE_FILL
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        current_row += 1

        # 헤더
        ws.cell(row=current_row, column=1, value="Genre").font = BOLD_FONT
        ws.cell(row=current_row, column=1).border = THIN_BORDER
        for j, l in enumerate(lengths, 2):
            cell = ws.cell(row=current_row, column=j, value=l)
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=6, value="Total").font = BOLD_FONT
        ws.cell(row=current_row, column=6).border = THIN_BORDER
        current_row += 1

        for genre in genres:
            ws.cell(row=current_row, column=1, value=genre).border = THIN_BORDER
            row_total = 0
            for j, lt in enumerate(lengths, 2):
                key = f"{genre}|{lt}"
                count = sample_counts.get(key, 0) if band == "700-900" else 0
                cell = ws.cell(row=current_row, column=j, value=count)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center')
                cell.fill = LIGHT_GREEN_FILL if count > 0 else LIGHT_RED_FILL
                row_total += count
            ws.cell(row=current_row, column=6, value=row_total).border = THIN_BORDER
            ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='center')
            current_row += 1

        current_row += 1  # 빈 행

    # 컬럼 너비
    ws.column_dimensions['A'].width = 18
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 12

    return ws


def create_question_bank_sheet(wb):
    """QUESTION_BANK 시트 생성"""
    ws = wb.create_sheet("QUESTION_BANK")

    headers = [
        "question_id", "source_text_id", "question_type", "question_type_name",
        "question_stem", "modified_passage",
        "choice_1", "choice_2", "choice_3", "choice_4", "choice_5",
        "correct_answer", "explanation", "difficulty", "created_date"
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    apply_header_style(ws, 1, 1, len(headers), fill=GREEN_FILL)
    ws.freeze_panes = "A2"

    # 컬럼 너비
    widths = {
        1: 30, 2: 22, 3: 14, 4: 16,
        5: 40, 6: 60,
        7: 30, 8: 30, 9: 30, 10: 30, 11: 30,
        12: 14, 13: 50, 14: 12, 15: 14
    }
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # 문항 유형 참조 테이블 (R열~)
    ws.cell(row=1, column=18, value="문항 유형 참조").font = TITLE_FONT
    ref_headers = ["코드", "유형명", "최소 Lexile", "최소 길이"]
    for i, h in enumerate(ref_headers):
        cell = ws.cell(row=2, column=18 + i, value=h)
        cell.font = BOLD_FONT
        cell.fill = LIGHT_BLUE_FILL
        cell.border = THIN_BORDER

    q_types = [
        ["MI-01", "글의 목적", 700, "Short"],
        ["MI-02", "주장/요지", 700, "Short"],
        ["MI-03", "주제", 700, "Short"],
        ["MI-04", "제목", 700, "Short"],
        ["DT-01", "내용 일치/불일치", 500, "Short"],
        ["IF-01", "함축 의미 추론", 900, "Short"],
        ["IF-02", "빈칸 추론 (구)", 900, "Short"],
        ["IF-03", "빈칸 추론 (문장)", 1100, "Medium"],
        ["ST-01", "무관한 문장", 900, "Medium"],
        ["ST-02", "문장 삽입", 900, "Medium"],
        ["ST-03", "글의 순서", 900, "Medium"],
        ["VG-01", "어법 판단", 700, "Short"],
        ["VG-02", "어휘 판단", 700, "Short"],
        ["SM-01", "요약문 완성", 1100, "Medium"],
    ]
    for i, row_data in enumerate(q_types, 3):
        for j, val in enumerate(row_data):
            cell = ws.cell(row=i, column=18 + j, value=val)
            cell.border = THIN_BORDER

    for col in range(18, 22):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # difficulty 드롭다운
    dv_diff = DataValidation(type="list", formula1='"easy,medium,hard,very_hard"', allow_blank=True)
    ws.add_data_validation(dv_diff)
    dv_diff.add("N2:N500")

    # question_type 드롭다운
    type_codes = ",".join([t[0] for t in q_types])
    dv_type = DataValidation(type="list", formula1=f'"{type_codes}"', allow_blank=True)
    ws.add_data_validation(dv_type)
    dv_type.add("C2:C500")

    return ws


def create_guide_sheet(wb):
    """사용 가이드 시트"""
    ws = wb.create_sheet("GUIDE")

    ws.cell(row=1, column=1, value="Lexile Reading Text DB — 사용 가이드").font = Font(bold=True, size=16)
    ws.merge_cells("A1:F1")

    guide = [
        "",
        "[ 시트 구성 ]",
        "  READING_TEXT_MASTER : 메인 지문 데이터베이스 (녹색=샘플, 노랑=빈 슬롯)",
        "  CONFIG              : 드롭다운 및 설정 데이터",
        "  COVERAGE_MATRIX     : Genre × Length 배치 현황 (녹색=있음, 빨강=없음)",
        "  QUESTION_BANK       : 수능형 문항 저장소",
        "",
        "[ 텍스트 추가 방법 ]",
        "  1. MASTER 시트의 노란색 빈 슬롯에서 topic 입력 → text_body에 지문 붙여넣기",
        "  2. 또는 Apps Script 메뉴 > AI Text Generation > Generate for Selected Row",
        "  3. text_body 입력 시 word_count, sentence_count 등이 자동 계산됨",
        "",
        "[ 장르 추가 방법 ]",
        "  1. Apps Script의 Code.gs에서 GENRE_CODES 객체에 새 장르 추가",
        "  2. setupConfigSheet_ 함수의 genres 배열에 추가",
        "  3. createCoverageMatrix 함수의 genres 배열에 추가",
        "  4. Setup Sheets 재실행",
        "",
        "[ 문항 생성 방법 ]",
        "  1. MASTER 시트에서 지문이 있는 행 선택",
        "  2. 메뉴 > CSAT Question Generation > Generate Question",
        "  3. 추천 문항 유형 중 선택 → QUESTION_BANK에 자동 저장",
        "",
        "[ Apps Script 설치 ]",
        "  1. Extensions > Apps Script 열기",
        "  2. apps-script/ 폴더의 .gs 파일들을 각각 추가",
        "  3. appsscript.json 설정 적용",
        "  4. 시트 새로고침 후 메뉴 확인",
        "",
        "[ 수능 모의고사 세트 구성 ]",
        "  QUESTION_BANK에서 유형별 필터 후 조합:",
        "  18번 MI-01 / 20번 MI-02 / 21번 MI-03 / 22번 MI-04",
        "  29번 VG-01 / 30번 VG-02 / 31-34번 IF-02,IF-03",
        "  35번 ST-01 / 36-37번 ST-03 / 38-39번 ST-02 / 40번 SM-01",
    ]

    for i, line in enumerate(guide, 3):
        cell = ws.cell(row=i, column=1, value=line)
        if line.startswith("[ "):
            cell.font = BOLD_FONT
            cell.fill = LIGHT_BLUE_FILL

    ws.column_dimensions['A'].width = 80

    return ws


# ==================== 메인 ====================

def main():
    wb = openpyxl.Workbook()

    print("1/5  READING_TEXT_MASTER 생성 중...")
    create_master_sheet(wb)

    print("2/5  CONFIG 생성 중...")
    create_config_sheet(wb)

    print("3/5  COVERAGE_MATRIX 생성 중...")
    create_coverage_sheet(wb)

    print("4/5  QUESTION_BANK 생성 중...")
    create_question_bank_sheet(wb)

    print("5/5  GUIDE 생성 중...")
    create_guide_sheet(wb)

    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "..", "Lexile_Reading_Text_DB.xlsx")
    output_path = os.path.normpath(output_path)

    wb.save(output_path)
    print(f"\n완료! 파일 저장 위치: {output_path}")
    print(f"  - READING_TEXT_MASTER: 샘플 {len(SAMPLE_DATA)}개 + 빈 슬롯 {len(TEMPLATE_SLOTS)}개")
    print(f"  - CONFIG: Lexile Bands, Genres, Length Types, CEFR, 용도, 연령 주제")
    print(f"  - COVERAGE_MATRIX: 7개 Band × 6 Genre × 4 Length")
    print(f"  - QUESTION_BANK: 14개 문항 유형 참조 포함")
    print(f"  - GUIDE: 사용 안내")


if __name__ == "__main__":
    main()
